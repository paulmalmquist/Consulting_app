"""Text extraction from PDF, DOCX, XLSX, and plain text files.

Demonstrates: structured + unstructured data pipeline for RAG ingestion.
All dependencies (pypdf, python-docx, openpyxl) are already in requirements.txt.
"""
from __future__ import annotations

import io


_IMAGE_MIMES = frozenset(
    ["image/png", "image/jpeg", "image/jpg", "image/webp", "image/gif"]
)
_IMAGE_EXTS = frozenset([".png", ".jpg", ".jpeg", ".webp", ".gif"])

# Sentinel returned when a file is an image and should not be text-indexed.
IMAGE_TRACK_SENTINEL = "__image_track__"


def extract_text(content: bytes, mime_type: str, filename: str = "") -> str:
    """Route bytes to the right extractor based on MIME type or extension.

    Returns plain text suitable for chunking and embedding.
    For image/* files returns IMAGE_TRACK_SENTINEL — callers must check
    and handle the image track separately (vision_only processing_mode).
    """
    fname = filename.lower()
    # Image short-circuit: return sentinel, never decode image bytes as text
    if mime_type in _IMAGE_MIMES or any(fname.endswith(ext) for ext in _IMAGE_EXTS):
        return IMAGE_TRACK_SENTINEL
    if mime_type == "application/pdf" or fname.endswith(".pdf"):
        return _extract_pdf(content)
    if mime_type in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ) or fname.endswith((".docx", ".doc")):
        return _extract_docx(content)
    if mime_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ) or fname.endswith((".xlsx", ".xls")):
        return _extract_xlsx(content)
    if fname.endswith(".csv"):
        return content.decode("utf-8", errors="replace")
    # Plain text / markdown fallback
    return content.decode("utf-8", errors="replace")


def _extract_pdf(content: bytes) -> str:
    """Extract text from PDF with page markers for citation traceability."""
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(content))
    pages = []
    for i, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if text.strip():
            pages.append(f"[Page {i}]\n{text.strip()}")
    return "\n\n".join(pages)


def _extract_docx(content: bytes) -> str:
    """Extract text from DOCX preserving paragraph structure."""
    from docx import Document

    doc = Document(io.BytesIO(content))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def _extract_xlsx(content: bytes) -> str:
    """Extract text from XLSX with sheet and row structure."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
        if rows:
            parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
    return "\n\n".join(parts)
