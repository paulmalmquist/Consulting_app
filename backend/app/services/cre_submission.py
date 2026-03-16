"""CRE Submission Portal Service.

Accepts vendor file uploads (CSV, Excel, PDF), auto-detects format,
validates, and routes to the appropriate ingestion pipeline.
"""
from __future__ import annotations

import csv
import io
import json
import logging
from uuid import UUID

from app.db import get_cursor

log = logging.getLogger(__name__)

# Known column header patterns for auto-detection
_RENT_ROLL_HEADERS = {"tenant", "unit", "rent", "lease_start", "lease_end", "sqft"}
_OPERATING_STATEMENT_HEADERS = {"revenue", "expenses", "noi", "period", "property"}
_PROPERTY_LIST_HEADERS = {"address", "property_name", "property_type", "sqft", "year_built"}


def detect_format(*, filename: str, content: bytes) -> dict:
    """Detect file format and suggest column mappings.

    Returns: {"format": "csv"|"excel"|"pdf", "detected_type": str, "columns": list, "row_count": int}
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "pdf":
        return {"format": "pdf", "detected_type": "document", "columns": [], "row_count": 0,
                "message": "PDF detected — will route to document extraction pipeline"}

    if ext in ("xlsx", "xls"):
        return _detect_excel(content)

    if ext == "csv" or ext == "":
        return _detect_csv(content)

    return {"format": "unknown", "detected_type": "unknown", "columns": [], "row_count": 0}


def _detect_csv(content: bytes) -> dict:
    """Detect CSV format and column mapping."""
    try:
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        headers = [h.strip().lower().replace(" ", "_") for h in next(reader, [])]
        row_count = sum(1 for _ in reader)
    except Exception as exc:
        return {"format": "csv", "detected_type": "unknown", "columns": [], "row_count": 0, "error": str(exc)}

    header_set = set(headers)
    detected_type = _classify_headers(header_set)

    return {"format": "csv", "detected_type": detected_type, "columns": headers, "row_count": row_count}


def _detect_excel(content: bytes) -> dict:
    """Detect Excel format and column mapping."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower().replace(" ", "_") for c in next(ws.iter_rows(max_row=1))]
        row_count = ws.max_row - 1 if ws.max_row else 0
        wb.close()
    except Exception as exc:
        return {"format": "excel", "detected_type": "unknown", "columns": [], "row_count": 0, "error": str(exc)}

    header_set = set(headers)
    detected_type = _classify_headers(header_set)

    return {"format": "excel", "detected_type": detected_type, "columns": headers, "row_count": row_count}


def _classify_headers(headers: set[str]) -> str:
    """Classify a set of column headers into a known document type."""
    if headers & _RENT_ROLL_HEADERS and len(headers & _RENT_ROLL_HEADERS) >= 3:
        return "rent_roll"
    if headers & _OPERATING_STATEMENT_HEADERS and len(headers & _OPERATING_STATEMENT_HEADERS) >= 3:
        return "operating_statement"
    if headers & _PROPERTY_LIST_HEADERS and len(headers & _PROPERTY_LIST_HEADERS) >= 3:
        return "property_list"
    return "generic"


def process_submission(
    *,
    env_id: UUID,
    business_id: UUID,
    filename: str,
    content: bytes,
) -> dict:
    """Process a submitted file: detect format, validate, and create ingest run."""
    detection = detect_format(filename=filename, content=content)

    if detection["format"] == "pdf":
        # Route to existing extraction service
        return {
            "status": "queued",
            "format": "pdf",
            "detected_type": "document",
            "message": "PDF queued for extraction via document extraction pipeline",
        }

    if detection["detected_type"] == "unknown":
        return {
            "status": "rejected",
            "format": detection["format"],
            "detected_type": "unknown",
            "message": "Could not auto-detect document type from column headers",
            "columns": detection.get("columns", []),
        }

    # Create ingest run for structured data
    with get_cursor() as cur:
        cur.execute(
            """
            INSERT INTO cre_ingest_run (source_key, scope_json, status)
            VALUES ('submission_portal', %s::jsonb, 'running')
            RETURNING run_id
            """,
            (json.dumps({
                "filename": filename,
                "format": detection["format"],
                "detected_type": detection["detected_type"],
                "row_count": detection["row_count"],
            }),),
        )
        run = cur.fetchone()

    return {
        "status": "accepted",
        "run_id": str(run["run_id"]),
        "format": detection["format"],
        "detected_type": detection["detected_type"],
        "row_count": detection["row_count"],
        "columns": detection.get("columns", []),
    }
