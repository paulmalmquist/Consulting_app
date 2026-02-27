"""Excel export service.

Generates .xlsx fund report with openpyxl.
Sheets: Fund Summary, LP Capital Accounts, Waterfall Breakdown,
        Loan Schedule, NOI Variance.
"""
from __future__ import annotations

import io
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font

from app.db import get_cursor


HEADER_FONT = Font(bold=True, size=11)
PCT_FMT = "0.00%"
MONEY_FMT = '#,##0.00'


def export_fund_report(
    *,
    env_id: str,
    business_id: UUID,
    fund_id: UUID,
    quarter: str,
) -> bytes:
    """Generate complete fund report as .xlsx bytes."""
    wb = Workbook()

    _build_fund_summary(wb, fund_id, quarter)
    _build_lp_capital_accounts(wb, fund_id, quarter)
    _build_waterfall_breakdown(wb, fund_id, quarter)
    _build_loan_schedules(wb, env_id, business_id, fund_id)
    _build_noi_variance(wb, env_id, business_id, quarter)

    # Remove default empty sheet if other sheets exist
    if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_fund_summary(wb: Workbook, fund_id: UUID, quarter: str) -> None:
    ws = wb.active
    ws.title = "Fund Summary"

    with get_cursor() as cur:
        cur.execute(
            """
            SELECT * FROM re_fund_metrics_qtr
            WHERE fund_id = %s AND quarter = %s
            ORDER BY created_at DESC LIMIT 1
            """,
            (str(fund_id), quarter),
        )
        metrics = cur.fetchone()

        cur.execute(
            "SELECT fund_name FROM repe_fund WHERE fund_id = %s",
            (str(fund_id),),
        )
        fund_row = cur.fetchone()

    fund_name = fund_row["fund_name"] if fund_row else str(fund_id)

    ws["A1"] = fund_name
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Quarter: {quarter}"

    if not metrics:
        ws["A4"] = "No metrics available for this quarter."
        return

    headers = ["Metric", "Value"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT

    kpis = [
        ("Gross IRR", metrics.get("gross_irr"), PCT_FMT),
        ("Net IRR", metrics.get("net_irr"), PCT_FMT),
        ("Gross TVPI", metrics.get("gross_tvpi"), "0.00x"),
        ("Net TVPI", metrics.get("net_tvpi"), "0.00x"),
        ("DPI", metrics.get("dpi"), "0.00x"),
        ("RVPI", metrics.get("rvpi"), "0.00x"),
        ("NAV", metrics.get("nav"), MONEY_FMT),
    ]

    for i, (label, value, fmt) in enumerate(kpis, 5):
        ws.cell(row=i, column=1, value=label)
        cell = ws.cell(row=i, column=2)
        if value is not None:
            cell.value = float(value)
            cell.number_format = fmt


def _build_lp_capital_accounts(wb: Workbook, fund_id: UUID, quarter: str) -> None:
    ws = wb.create_sheet("LP Capital Accounts")

    with get_cursor() as cur:
        cur.execute(
            """
            SELECT s.*, p.partner_name, p.partner_type
            FROM re_capital_account_snapshot s
            JOIN re_partner p ON p.id = s.partner_id
            WHERE s.fund_id = %s AND s.quarter = %s
            ORDER BY p.partner_type, p.partner_name
            """,
            (str(fund_id), quarter),
        )
        rows = cur.fetchall()

    headers = [
        "Partner", "Type", "Committed", "Contributed", "Distributed",
        "NAV Share", "DPI", "RVPI", "TVPI", "Carry Allocation",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT

    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r.get("partner_name", ""))
        ws.cell(row=i, column=2, value=r.get("partner_type", ""))
        for col, key in enumerate(
            ["committed", "contributed", "distributed", "nav_share"], 3
        ):
            cell = ws.cell(row=i, column=col)
            cell.value = float(r[key]) if r.get(key) else 0
            cell.number_format = MONEY_FMT
        for col, key in enumerate(["dpi", "rvpi", "tvpi"], 7):
            cell = ws.cell(row=i, column=col)
            cell.value = float(r[key]) if r.get(key) else 0
            cell.number_format = "0.00x"
        cell = ws.cell(row=i, column=10)
        cell.value = float(r["carry_allocation"]) if r.get("carry_allocation") else 0
        cell.number_format = MONEY_FMT


def _build_waterfall_breakdown(wb: Workbook, fund_id: UUID, quarter: str) -> None:
    ws = wb.create_sheet("Waterfall Breakdown")

    with get_cursor() as cur:
        cur.execute(
            """
            SELECT wr.id FROM re_waterfall_run wr
            WHERE wr.fund_id = %s AND wr.quarter = %s
            ORDER BY wr.created_at DESC LIMIT 1
            """,
            (str(fund_id), quarter),
        )
        run = cur.fetchone()
        if not run:
            ws["A1"] = "No waterfall run for this quarter."
            return

        cur.execute(
            """
            SELECT wrr.tier_name, wrr.amount, p.partner_name, p.partner_type
            FROM re_waterfall_run_result wrr
            JOIN re_partner p ON p.id = wrr.partner_id
            WHERE wrr.run_id = %s
            ORDER BY wrr.tier_name, p.partner_type, p.partner_name
            """,
            (str(run["id"]),),
        )
        rows = cur.fetchall()

    headers = ["Tier", "Partner", "Type", "Amount"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT

    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r.get("tier_name", ""))
        ws.cell(row=i, column=2, value=r.get("partner_name", ""))
        ws.cell(row=i, column=3, value=r.get("partner_type", ""))
        cell = ws.cell(row=i, column=4)
        cell.value = float(r["amount"]) if r.get("amount") else 0
        cell.number_format = MONEY_FMT


def _build_loan_schedules(
    wb: Workbook, env_id: str, business_id: UUID, fund_id: UUID
) -> None:
    ws = wb.create_sheet("Loan Schedules")

    with get_cursor() as cur:
        cur.execute(
            """
            SELECT l.id, l.loan_name
            FROM re_loan l
            WHERE l.env_id = %s AND l.business_id = %s AND l.fund_id = %s
            ORDER BY l.loan_name
            """,
            (env_id, str(business_id), str(fund_id)),
        )
        loans = cur.fetchall()

        if not loans:
            ws["A1"] = "No loans in this fund."
            return

        row_num = 1
        for loan in loans:
            ws.cell(row=row_num, column=1, value=loan["loan_name"])
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
            row_num += 1

            headers = [
                "Period", "Beginning Balance", "Principal",
                "Interest", "Total Payment", "Ending Balance",
            ]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=h)
                cell.font = HEADER_FONT
            row_num += 1

            cur.execute(
                """
                SELECT * FROM re_loan_amortization_schedule
                WHERE loan_id = %s ORDER BY period_number
                """,
                (str(loan["id"]),),
            )
            schedule = cur.fetchall()
            for s in schedule:
                ws.cell(row=row_num, column=1, value=s["period_number"])
                for col, key in enumerate(
                    ["beginning_balance", "scheduled_principal",
                     "interest_payment", "total_payment", "ending_balance"], 2
                ):
                    cell = ws.cell(row=row_num, column=col)
                    cell.value = float(s[key]) if s.get(key) else 0
                    cell.number_format = MONEY_FMT
                row_num += 1

            row_num += 1  # blank row between loans


def _build_noi_variance(
    wb: Workbook, env_id: str, business_id: UUID, quarter: str
) -> None:
    ws = wb.create_sheet("NOI Variance")

    with get_cursor() as cur:
        cur.execute(
            """
            SELECT v.*, a.asset_name
            FROM re_asset_variance_qtr v
            JOIN repe_asset a ON a.asset_id = v.asset_id
            WHERE v.env_id = %s AND v.business_id = %s AND v.quarter = %s
            ORDER BY a.asset_name
            """,
            (env_id, str(business_id), quarter),
        )
        rows = cur.fetchall()

    headers = [
        "Asset", "Line Code", "Actual", "Plan", "Variance $", "Variance %",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT

    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r.get("asset_name", ""))
        ws.cell(row=i, column=2, value=r.get("line_code", ""))
        for col, key in enumerate(["actual", "plan", "variance_dollar"], 3):
            cell = ws.cell(row=i, column=col)
            cell.value = float(r[key]) if r.get(key) else 0
            cell.number_format = MONEY_FMT
        cell = ws.cell(row=i, column=6)
        cell.value = float(r["variance_pct"]) if r.get("variance_pct") else 0
        cell.number_format = PCT_FMT
