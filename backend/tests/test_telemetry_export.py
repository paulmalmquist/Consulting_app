"""Tests for the read-only telemetry XLSX export route (Phase 8 / PR 8A).

Covers: allowed dataset, denied dataset, row-limit bound, empty result, and the
generated file shape. No DB — the underlying read service is monkeypatched, which also
proves the route never builds SQL from request input (it only calls the allowlisted
service)."""
import io
import os
import sys
import types
from uuid import UUID

os.environ.setdefault("DATABASE_URL", "postgresql://test:test@localhost:5432/test")
os.environ.setdefault("SUPABASE_URL", "https://test.supabase.co")
os.environ.setdefault("SUPABASE_SERVICE_ROLE_KEY", "test-key")
os.environ.setdefault("_BM_SKIP_DB_CHECK", "1")

if "psycopg" not in sys.modules:
    psycopg_stub = types.ModuleType("psycopg")
    psycopg_stub.connect = lambda *a, **k: None
    psycopg_stub.Connection = object
    psycopg_stub.rows = types.SimpleNamespace(dict_row=None)
    sys.modules["psycopg"] = psycopg_stub
if "dotenv" not in sys.modules:
    dotenv_stub = types.ModuleType("dotenv")
    dotenv_stub.load_dotenv = lambda *a, **k: None
    sys.modules["dotenv"] = dotenv_stub

from fastapi.responses import JSONResponse, Response  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.routes import telemetry_export as te  # noqa: E402

BIZ = UUID("00000000-0000-0000-0000-000000000002")
COLS = te.TELEMETRY_EXPORT_DATASETS["model_runs"].columns


def _fake_models(rows, null_reason=None):
    return lambda **kw: {"models": rows, "null_reason": null_reason}


def test_workbook_shape_header_then_rows():
    rows = [{"model_name": "cnn-lstm", "model_kind": "rul", "metrics": {"rmse": 17.33}}]
    data = te._workbook_bytes(COLS, rows)
    ws = load_workbook(io.BytesIO(data)).active
    assert [c.value for c in ws[1]] == list(COLS)            # header row = fixed columns
    assert ws.cell(row=2, column=1).value == "cnn-lstm"       # data row
    # jsonb cell is flattened to a string, value preserved (rmse 17.33 present)
    metrics_idx = list(COLS).index("metrics") + 1
    assert "17.33" in str(ws.cell(row=2, column=metrics_idx).value)


def test_denied_dataset_is_404_with_null_reason():
    resp = te.export_dataset_xlsx("totally_made_up", env_id="telemetry-demo", business_id=BIZ, limit=None)
    assert isinstance(resp, JSONResponse)
    assert resp.status_code == 404
    import json
    body = json.loads(bytes(resp.body))
    assert body["null_reason"] == "export_dataset_not_allowed"
    assert "model_runs" in body["allowed"]


def test_allowed_dataset_streams_xlsx_with_row_header(monkeypatch):
    monkeypatch.setattr(te.telemetry_serving, "model_performance",
                        _fake_models([{"model_name": "cnn", "model_kind": "rul"}]))
    resp = te.export_dataset_xlsx("model_runs", env_id="telemetry-demo", business_id=BIZ, limit=None)
    assert isinstance(resp, Response)
    assert resp.media_type == te.XLSX_MEDIA_TYPE
    assert resp.headers["X-Telemetry-Export-Rows"] == "1"
    assert resp.headers["X-Telemetry-Export-Source-Kind"] == "live-rows"
    assert "attachment" in resp.headers["Content-Disposition"]
    ws = load_workbook(io.BytesIO(resp.body)).active   # body is the real xlsx
    assert ws.cell(row=2, column=1).value == "cnn"


def test_row_limit_is_bounded(monkeypatch):
    many = [{"model_name": f"m{i}", "model_kind": "rul"} for i in range(10)]
    monkeypatch.setattr(te.telemetry_serving, "model_performance", _fake_models(many))
    resp = te.export_dataset_xlsx("model_runs", env_id="telemetry-demo", business_id=BIZ, limit=3)
    assert resp.headers["X-Telemetry-Export-Rows"] == "3"
    assert resp.headers["X-Telemetry-Export-Limit"] == "3"


def test_max_limit_caps_request(monkeypatch):
    monkeypatch.setattr(te.telemetry_serving, "model_performance", _fake_models([]))
    resp = te.export_dataset_xlsx("model_runs", env_id="telemetry-demo", business_id=BIZ, limit=999999)
    # eff_limit is capped at the dataset max_limit regardless of the request
    assert int(resp.headers["X-Telemetry-Export-Limit"]) == te.TELEMETRY_EXPORT_DATASETS["model_runs"].max_limit


def test_empty_result_is_valid_workbook_with_null_reason(monkeypatch):
    monkeypatch.setattr(te.telemetry_serving, "model_performance",
                        _fake_models([], null_reason="model_not_promoted"))
    resp = te.export_dataset_xlsx("model_runs", env_id="telemetry-demo", business_id=BIZ, limit=None)
    assert resp.headers["X-Telemetry-Export-Rows"] == "0"
    assert resp.headers["X-Telemetry-Null-Reason"] == "model_not_promoted"
    ws = load_workbook(io.BytesIO(resp.body)).active
    assert [c.value for c in ws[1]] == list(COLS)   # header-only, still a valid file
    assert ws.max_row == 1


def test_list_datasets_exposes_allowlist_no_data():
    out = te.list_export_datasets()
    names = {d["dataset"] for d in out["datasets"]}
    assert "model_runs" in names
    assert "anomaly_events" in names
    assert all("source_kind" in d for d in out["datasets"])


# ── anomaly_events dataset (8D-3): reuses the read-only stream_live scoped path ──────────────────
def _fake_stream_live(events, null_reason=None):
    def _f(*, env_id, business_id, **kw):
        return {"events": events, "null_reason": null_reason, "channels": [], "source": "db_tail"}
    return _f


def _stub_stream_etl(monkeypatch, events, null_reason=None):
    stub = types.ModuleType("app.services.telemetry_stream_etl")
    stub.stream_live = _fake_stream_live(events, null_reason)
    monkeypatch.setitem(sys.modules, "app.services.telemetry_stream_etl", stub)


ANOMALY_COLS = list(te.TELEMETRY_EXPORT_DATASETS["anomaly_events"].columns)


def test_anomaly_events_streams_xlsx_with_fixed_columns(monkeypatch):
    _stub_stream_etl(monkeypatch, [
        {"channel_name": "USLAB000058", "start_t": 1, "end_t": 2, "anomaly_class": "point", "confidence": 0.9},
    ])
    resp = te.export_dataset_xlsx("anomaly_events", env_id="telemetry-demo", business_id=BIZ, limit=None)
    assert isinstance(resp, Response)
    assert resp.headers["X-Telemetry-Export-Rows"] == "1"
    assert resp.headers["X-Telemetry-Export-Source"] == "tel_anomaly_events"
    assert resp.headers["X-Telemetry-Export-Source-Kind"] == "live-rows"
    ws = load_workbook(io.BytesIO(resp.body)).active
    assert [c.value for c in ws[1]] == ANOMALY_COLS
    assert ws.cell(row=2, column=1).value == "USLAB000058"


def test_anomaly_events_row_limit_capped(monkeypatch):
    many = [{"channel_name": f"c{i}", "start_t": i, "end_t": i + 1, "anomaly_class": "point", "confidence": 0.5}
            for i in range(10)]
    _stub_stream_etl(monkeypatch, many)
    resp = te.export_dataset_xlsx("anomaly_events", env_id="telemetry-demo", business_id=BIZ, limit=3)
    assert resp.headers["X-Telemetry-Export-Rows"] == "3"


def test_anomaly_events_empty_is_header_only_with_null_reason(monkeypatch):
    _stub_stream_etl(monkeypatch, [], null_reason="stream_worker_disabled")
    resp = te.export_dataset_xlsx("anomaly_events", env_id="telemetry-demo", business_id=BIZ, limit=None)
    assert resp.headers["X-Telemetry-Export-Rows"] == "0"
    assert resp.headers["X-Telemetry-Null-Reason"] == "stream_worker_disabled"
    ws = load_workbook(io.BytesIO(resp.body)).active
    assert [c.value for c in ws[1]] == ANOMALY_COLS
    assert ws.max_row == 1


def test_anomaly_events_empty_no_stream_reason_uses_specific_null_reason(monkeypatch):
    _stub_stream_etl(monkeypatch, [], null_reason=None)
    resp = te.export_dataset_xlsx("anomaly_events", env_id="telemetry-demo", business_id=BIZ, limit=None)
    assert resp.headers["X-Telemetry-Null-Reason"] == "no_live_anomaly_events"
