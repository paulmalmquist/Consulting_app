from __future__ import annotations

import argparse
import csv
import json
import sys
from copy import copy
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
BACKEND_ROOT = ROOT / "backend"
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.services import operator_property_ops as property_ops  # noqa: E402


PURPLE = "35146B"
VIOLET = "5430C0"
MINT = "C6F4DF"
CREAM = "FBFAF7"
LINE = "DDD8EA"
INK = "241437"
MUTED = "6F6590"
WARN = "FFF1CC"
RISK = "FAD4D8"
GOOD = "DDF7E8"


def _csv_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _format_value(value: Any) -> Any:
    if isinstance(value, (list, tuple, set)):
        return "; ".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, sort_keys=True)
    return value


def _write_table(ws, name: str, rows: Iterable[dict[str, Any]], *, start_row: int = 1) -> int:
    rows = list(rows)
    if not rows:
        ws.cell(start_row, 1, "No rows available")
        return start_row
    headers = list(rows[0].keys())
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col_idx, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=PURPLE)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, _format_value(row.get(header)))
    end_row = start_row + len(rows)
    end_col = len(headers)
    ref = f"A{start_row}:{ws.cell(end_row, end_col).coordinate}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = f"A{start_row + 1}"
    for idx, header in enumerate(headers, start=1):
        width = min(max(len(str(header)) + 6, 14), 34)
        ws.column_dimensions[ws.cell(start_row, idx).column_letter].width = width
    return end_row


def _style_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color=LINE)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            border = copy(cell.border)
            border.bottom = thin
            cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.0"


def _summary_sheet(wb: Workbook, metadata: dict[str, Any], ml_metrics: dict[str, Any], output_path: Path) -> None:
    ws = wb.active
    ws.title = "Read Me"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "HappyCo Property Ops Intelligence Kit"
    ws["A1"].font = Font(size=20, bold=True, color=PURPLE)
    ws["A2"] = "Synthetic demo workbook for Head of Data proof-of-work. Not HappyCo production data."
    ws["A2"].font = Font(size=11, color=INK)
    rows = [
        ("Fixture version", metadata.get("fixture_version")),
        ("Generated at", metadata.get("generated_at")),
        ("Demo mode", str(metadata.get("demo_mode"))),
        ("Data source", metadata.get("data_source")),
        ("Caveat", metadata.get("caveat")),
        ("Workbook path", str(output_path)),
        ("ML backend", ml_metrics.get("training_backend", "not_available")),
        ("ML honesty warning", ml_metrics.get("honesty_warning", "ML artifacts not available.")),
    ]
    for idx, (label, value) in enumerate(rows, start=4):
        ws.cell(idx, 1, label).font = Font(bold=True, color=PURPLE)
        ws.cell(idx, 2, value)
    ws["A14"] = "Workbook Contents"
    ws["A14"].font = Font(size=14, bold=True, color=PURPLE)
    contents = [
        ("Canonical Entities", "=COUNTA('Canonical Entities'!A:A)-1"),
        ("Property Benchmarks", "=COUNTA('Property Benchmarks'!A:A)-1"),
        ("Inspection Risk", "=COUNTA('Inspection Risk'!A:A)-1"),
        ("Work Order Aging", "=COUNTA('Work Order Aging'!A:A)-1"),
        ("Vendor Performance", "=COUNTA('Vendor Performance'!A:A)-1"),
        ("AI Recommendations", "=COUNTA('AI Recommendations'!A:A)-1"),
        ("ML Feature Table", "=COUNTA('ML Feature Table'!A:A)-1"),
        ("ML Predictions", "=COUNTA('ML Predictions'!A:A)-1"),
    ]
    ws.append([])
    ws.append(["Tab", "Row count formula"])
    for tab, formula in contents:
        ws.append([tab, formula])
    ws["D14"] = "Model Metrics"
    ws["D14"].font = Font(size=14, bold=True, color=PURPLE)
    metric_rows = [
        ("Accuracy", ml_metrics.get("accuracy")),
        ("ROC AUC", ml_metrics.get("roc_auc")),
        ("Rows", ml_metrics.get("total_rows")),
        ("Positive rows", ml_metrics.get("positive_rows")),
        ("Model version", ml_metrics.get("model_version")),
    ]
    for idx, (label, value) in enumerate(metric_rows, start=15):
        ws.cell(idx, 4, label).font = Font(bold=True, color=PURPLE)
        ws.cell(idx, 5, value)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 92
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 24
    for cell in ws["A1:E1"][0]:
        cell.fill = PatternFill("solid", fgColor=MINT)


def _canonical_rows(dataset: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    specs = [
        ("operator", "operator_id", dataset["operators"]),
        ("property", "property_id", dataset["properties"]),
        ("building", "building_id", dataset["buildings"]),
        ("unit", "unit_id", dataset["units"]),
        ("vendor", "vendor_id", dataset["vendors"]),
    ]
    for entity_type, id_key, entities in specs:
        for row in entities:
            rows.append(
                {
                    "entity_type": entity_type,
                    "entity_id": row[id_key],
                    "name": row.get("name") or row.get("canonical_label") or row.get("unit_code"),
                    "property_id": row.get("property_id"),
                    "business_id": row.get("business_id"),
                    "source_systems": row.get("source_systems"),
                    "peer_group": row.get("peer_group"),
                    "synthetic_demo": True,
                }
            )
    return rows


def _inspection_rows(dataset: dict[str, Any]) -> list[dict[str, Any]]:
    inspection_by_id = {row["inspection_id"]: row for row in dataset["inspections"]}
    property_by_id = {row["property_id"]: row for row in dataset["properties"]}
    rows = []
    for finding in dataset["findings"]:
        inspection = inspection_by_id[finding["inspection_id"]]
        prop = property_by_id[finding["property_id"]]
        rows.append(
            {
                "finding_id": finding["finding_id"],
                "property_id": finding["property_id"],
                "property_name": prop["name"],
                "building_id": finding["building_id"],
                "inspection_id": finding["inspection_id"],
                "inspection_date": inspection["inspection_date"],
                "category": finding["category"],
                "severity_score": finding["severity_score"],
                "requires_work_order": finding["requires_work_order"],
                "deficiency_code": finding["deficiency_code"],
                "description": finding["description"],
            }
        )
    return rows


def _work_order_rows(dataset: dict[str, Any]) -> list[dict[str, Any]]:
    property_by_id = {row["property_id"]: row for row in dataset["properties"]}
    vendor_by_id = {row["vendor_id"]: row for row in dataset["vendors"]}
    rows = []
    for work_order in dataset["work_orders"]:
        rows.append(
            {
                "work_order_id": work_order["work_order_id"],
                "property_id": work_order["property_id"],
                "property_name": property_by_id[work_order["property_id"]]["name"],
                "building_id": work_order["building_id"],
                "unit_id": work_order["unit_id"],
                "category": work_order["category"],
                "status": work_order["status"],
                "priority": work_order["priority"],
                "opened_date": work_order["opened_date"],
                "closed_date": work_order["closed_date"],
                "aging_days": work_order["aging_days"],
                "reopened": work_order["reopened"],
                "repeat_flag": work_order["repeat_flag"],
                "vendor_id": work_order["vendor_id"],
                "vendor_name": vendor_by_id[work_order["vendor_id"]]["name"],
                "aging_risk_score": None,
            }
        )
    return rows


def _recommendation_rows(recommendations_payload: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for row in recommendations_payload["recommendations"]:
        rows.append(
            {
                "recommendation_id": row["recommendation_id"],
                "property_id": row["property_id"],
                "title": row["title"],
                "risk_level": row["risk_level"],
                "confidence": row["confidence"],
                "recommended_action": row["recommended_action"],
                "human_review_required": row["human_review_required"],
                "recommendation": row["recommendation"],
                "evidence_ids": row["evidence_ids"],
                "ml_status": row.get("ml_status"),
                "ml_risk_score": row.get("ml_risk_score"),
                "risk_band": row.get("risk_band"),
                "top_model_drivers": row.get("top_model_drivers"),
                "caveats": row.get("caveats"),
            }
        )
    return rows


def build_workbook(output_path: Path, ml_root: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    dataset = property_ops.get_dataset()
    metadata = dataset["metadata"]
    ml_metrics = _json(ml_root / "model_metrics.json")
    feature_rows = _csv_rows(ml_root / "feature_table.csv") or property_ops.ml_feature_rows()
    prediction_rows = _csv_rows(ml_root / "predictions.csv")
    recommendation_payload = property_ops.recommendations_with_ml()

    wb = Workbook()
    _summary_sheet(wb, metadata, ml_metrics, output_path)

    sheet_rows = [
        ("Canonical Entities", "CanonicalEntities", _canonical_rows(dataset)),
        ("Property Benchmarks", "PropertyBenchmarks", dataset["benchmarks"]),
        ("Inspection Risk", "InspectionRisk", _inspection_rows(dataset)),
        ("Work Order Aging", "WorkOrderAging", _work_order_rows(dataset)),
        ("Vendor Performance", "VendorPerformance", dataset["vendor_performance"]),
        ("AI Recommendations", "AIRecommendations", _recommendation_rows(recommendation_payload)),
        ("ML Feature Table", "MLFeatureTable", feature_rows),
        ("ML Predictions", "MLPredictions", prediction_rows),
        ("Model Metrics", "ModelMetrics", [{"metric": key, "value": _format_value(value)} for key, value in ml_metrics.items()]),
    ]

    for sheet_name, table_name, rows in sheet_rows:
        ws = wb.create_sheet(sheet_name)
        end_row = _write_table(ws, table_name, rows)
        _style_sheet(ws)
        if sheet_name == "Work Order Aging" and end_row > 1:
            headers = [ws.cell(1, idx).value for idx in range(1, ws.max_column + 1)]
            score_col = headers.index("aging_risk_score") + 1
            aging_col = headers.index("aging_days") + 1
            status_col = headers.index("status") + 1
            repeat_col = headers.index("repeat_flag") + 1
            for row_idx in range(2, end_row + 1):
                aging_ref = ws.cell(row_idx, aging_col).coordinate
                status_ref = ws.cell(row_idx, status_col).coordinate
                repeat_ref = ws.cell(row_idx, repeat_col).coordinate
                ws.cell(row_idx, score_col, f'=IF({aging_ref}="","",{aging_ref}+IF({status_ref}="reopened",10,0)+IF({repeat_ref}=TRUE,6,0))')
            ws.conditional_formatting.add(
                f"{ws.cell(2, score_col).coordinate}:{ws.cell(end_row, score_col).coordinate}",
                ColorScaleRule(start_type="min", start_color=GOOD, mid_type="percentile", mid_value=50, mid_color=WARN, end_type="max", end_color=RISK),
            )
        if sheet_name == "Property Benchmarks" and end_row > 2:
            chart = BarChart()
            chart.title = "Repeat Work Order Rate by Property"
            chart.y_axis.title = "Repeat Rate %"
            chart.x_axis.title = "Property"
            headers = [ws.cell(1, idx).value for idx in range(1, ws.max_column + 1)]
            name_col = headers.index("property_name") + 1
            rate_col = headers.index("repeat_work_order_rate") + 1
            data = Reference(ws, min_col=rate_col, min_row=1, max_row=end_row)
            cats = Reference(ws, min_col=name_col, min_row=2, max_row=end_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 8
            chart.width = 16
            ws.add_chart(chart, "P2")

    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = PURPLE
        for row in ws.iter_rows():
            for cell in row:
                if cell.row == 1 and ws.title != "Read Me":
                    cell.font = Font(bold=True, color="FFFFFF")
                elif cell.value is not None:
                    cell.font = Font(color=INK)
        ws.freeze_panes = ws.freeze_panes or "A2"

    wb.save(output_path)
    validate_workbook(output_path)
    return output_path


def validate_workbook(path: Path) -> None:
    wb = load_workbook(path, data_only=False)
    required = {
        "Read Me",
        "Canonical Entities",
        "Property Benchmarks",
        "Inspection Risk",
        "Work Order Aging",
        "Vendor Performance",
        "AI Recommendations",
        "ML Feature Table",
        "ML Predictions",
        "Model Metrics",
    }
    missing = required.difference(wb.sheetnames)
    if missing:
        raise RuntimeError(f"Workbook missing sheets: {sorted(missing)}")
    if wb["Canonical Entities"].max_row < 100:
        raise RuntimeError("Canonical Entities sheet is unexpectedly sparse.")
    if wb["Work Order Aging"]["P2"].value is None:
        raise RuntimeError("Work Order Aging formula column is missing.")


def main() -> int:
    parser = argparse.ArgumentParser(description="Build the HappyCo Property Ops Excel workbook artifact.")
    parser.add_argument("--out", type=Path, default=ROOT / "artifacts" / "happyco" / "excel" / "HappyCo_Property_Ops_Model.xlsx")
    parser.add_argument("--ml-root", type=Path, default=ROOT / "artifacts" / "happyco" / "ml")
    args = parser.parse_args()
    output = build_workbook(args.out, args.ml_root)
    print(f"Wrote {output}")
    print("Validated sheets: Read Me, Canonical Entities, Property Benchmarks, Inspection Risk, Work Order Aging, Vendor Performance, AI Recommendations, ML Feature Table, ML Predictions, Model Metrics")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
